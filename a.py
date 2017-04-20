'''
	Author: Erik Zorn - Wallentin.
	Last Edit: April. 4 / 2017.
	
	Report is in this directory called "ezornwal_report.docx or pdf version".
	
	Results: Provided in the same directory this README is in. 
	For each specific "pass" results they have their own folder.
	For the FULL results of first pass see: "firstpass_matches.xlsx" or "firstpass_nomatches.xlsx" in First Pass Results folder.
	For the FULL results of second pass see: "secondpass_matches.xlsx" or "secondpass_nomatches.xlsx" in Second Pass Results folder.
	For the FULL results of possible matches see: "possible_matches.xlsx" or "possible_nomatches.xlsx" in Possible Matches Results folder.
	
	My code is all in the Python file called "a.py" which contains more info below:
	*** IMPORTANT READ *** 
	The code will probably not compile on your computer because of the import settings, please change the import settings based on your computer!
	Which is probably not the same way yours is setup!
	
	Python libraries being used:
	Openpyxl, csv, re, sys, timeit, datetime, time
	
	*** IMPORTANT READ *** 
		
	This program was created for CIS*4910 record linkage project with Luiza Antonie.
		
	The program will attempt to perform record linkage with several passes.
	1. First pass requirements: Hard coded district range see a.py and change lower and upper range in menu option "1", also needs files "1871full.txt" and "industrial.csv" in project directory to work.
	2. Second pass requirements: Hard coded district range see a.py and change lower and upper range in menu option "2", also needs files "1871full.txt" and "firstpass_nomatches.csv" in project directory to work.
	3. Possible matches requirements: Hard coded district range see a.py and change lower and upper range in menu option "3", also needs files "1871full.txt" and "secondpass_nomatches.csv" in project directory to work.
	
	The program contains error checking in the menu.
		
	It starts off by waiting for user input with a menu displayed to the user.
	Menu:
		1) First pass
		2) Second pass
		3) Find Possible Matches
		4) Quit the program (q)
	Choosing an option from the menu will allow you to do a specific task and you will need to wait for it to complete.
	Once it gives you the result from the task it will return you to the menu.
	
	Example Use:
		Choose menu option 3.
		It has hard coded file and district range so see a.py and menu option "3" to change.
		NOTE: Need "1871full.txt" and "firstpass_nomatches.csv" in project directory to work.
		Wait a for districts filtering and record linkage to finish.
		New excel file called "possible_matches.xlsx" will be created based on district range to be viewed.
		Choose a new menu option to do any more tasks.
		
	Limitations:
	- For safe record linkage and not run into memory issues, only compile from a range of 50 districts at once, do not do 1-206 (max district range) or it will crash because of Openpyxl library.
	- Does not handle matching of people from different districts.
	- Everything is hard coded, so not user friendly and programmer will need to change district range and files inside of the a.py file in the menu options as described above.
	
	References:
	https://openpyxl.readthedocs.io/en/default/
	http://stackoverflow.com/questions/2972212/creating-an-empty-list-in-python
	http://stackoverflow.com/questions/53513/best-way-to-check-if-a-list-is-empty
	https://docs.python.org/2/library/re.html
	https://docs.python.org/2/library/time.html
	https://docs.python.org/2/library/sys.html
	https://docs.python.org/2/library/datetime.html
'''

import sys, timeit, datetime
import time

from openpyxl import load_workbook
from openpyxl.utils import quote_sheetname
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

import csv
import re

'''
	Purpose: The main menu that is displayed to the user.
	Parameters: NONE.
	Return: NONE.
'''
def menu():
	print("\n\nPlease choose one of the following options using (1,2,3,4 or q):")
	print("1) First Pass")
	print("2) Second Pass")
	print("3) Find Possible Matches")
	print("4) Quit the program (q)")

'''
	Purpose: Gets the file given by the user using also the extension.
	Parameters: NONE.
	Return: file.
'''
def getFilename():
	filename = raw_input('Enter a file name with extension: ')
	return filename
	
'''
	Purpose: Gets the CSV file and all the data.
	Parameters: filename (the file name for the csv file), delimiter (delimiter used to separate the cells).
	Return: csv file data.
'''
def GetCSVFile(filename, delimiter):
	rows = []
	with open(filename, 'r') as f:
		reader = csv.reader(f, dialect ='excel', delimiter = delimiter)					
		#Iterate through csv file.
		for rowIndex, row in enumerate(reader):
			#print ("Row Index: %s" % rowIndex)
			rowData = []
			
			# Go through entire row.
			for columnIndex, cell in enumerate(row):
				#print ("Column Index: %s" % columnIndex)
				# Store the row.
				rowData.append(cell)					
				#print (cell)							
			#print ("")
			rows.append(rowData)
	return rows
	
'''
	Purpose: Gets the total row counts of the CSV file.
	Parameters: csvFile (the file name for the csv file).
	Return: row of data.
'''
def GetCSVTotalRows(csvFile):
	rowCount = 0
	# Count Total rows.
	for row in xrange(len(csvFile)-1, len(csvFile)):
		rowCount = int(row)
		#print (rowCount)
	return rowCount
	
'''
	Purpose: Gets first row of data from CSV file.
	Parameters: csvFile (the file name for the csv file).
	Return: row of data.
'''
def GetCSVFirstRow(csvFile):
	firstRow = []
	# Record the first row headers.
	for row in csvFile[:1]:
		#print (row)
		for column in row[:]:
			#print (column)
			firstRow.append(str(column))
	return firstRow

'''
	Purpose: Gets all the data from the district number.
	Parameters: filename (the file name for the csv file), delimiter (delimiter to parse the data), index (index number of cell data being passed in), district (district number being parsed).
	Return: the rows of data.
'''
def DistrictFilter(filename, delimiter, index, district):
	filteredRows = []
	with open(filename, 'r') as f:
		reader = csv.reader(f, dialect ='excel', delimiter = delimiter)					
		#Iterate through csv file.
		for rowIndex, row in enumerate(reader):
			#print ("Row Index: %s" % rowIndex)
			rowData = []
			
			# Record the first row headers.
			if (rowIndex == 0):
				# Go through entire row.
				for columnIndex, cell in enumerate(row):
					#print ("Column Index: %s" % columnIndex)
					 
					# Store the row.
					#cellStr = unicode(cell, errors = 'replace')
					#rowData.append(cellStr)
					rowData.append(cell)
					#print (cell)							
				#print ("")
				filteredRows.append(rowData)
			# Ignore header row.
			if (rowIndex != 0):					
				if (row[index] != ""):
					distValue = int(row[index])
					#print ("Dist: %s" % distValue)
					if (distValue == district):
						# Entire row is in the dist range.
						# Go through entire row.
						for columnIndex, cell in enumerate(row):
							#print ("Column Index: %s" % columnIndex)
							#cellStr = unicode(cell, errors = 'ignore')
							#rowData.append(cellStr)
							rowData.append(cell)
							#print (cell)
						# Collect the row data.
						filteredRows.append(rowData)
	return filteredRows
	
'''
	Purpose: Gets all the data from the district number.
	Parameters: index (index number of cell data being passed in), district (district number being parsed), record (record to match to), csvFile (the file name for the csv file).
	Return: matched row.
'''
def FindDuplicate(index, record, csvFile):
	foundMatch = False
	matchedRowData = []
	
	tempFile = csvFile
		
	for rowIndex, row in enumerate(tempFile):
		#print ("\nRow: %s" % row)	
		if (rowIndex != 0):
			#print ("\nIndex: %s" % rowIndex)	
			proprior = row[8]
			#print ("\nProprior: %s" % proprior)
			#print ("\nRecord: %s" % record)
			proprior = proprior.split()
			#print (record)
			#record = record.split()
			#print ("\nAFTER Proprior: %s" % proprior)
			#print ("\nAFTER Record: %s" % record)
			#print ("\nIndex: %s" % rowIndex)		
			#print ("Proprior: %s" % proprior)
			
			# Don't look if any of the fields are empty.
			if proprior and record:
				if (proprior == record and index != rowIndex):
					foundMatch = True
		
				if (foundMatch):
					matchedRowData = []
					for columnIndex, cell in enumerate(row):
						matchedRowData.append(cell)
						
					#print ("Found Match: %s" % matchedRowData)
					#print ("\nIndex: %s" % rowIndex)		
					#print ("FOUND Proprior: %s" % proprior)
					# Remove found record.
					csvFile.pop(rowIndex)
					
					return matchedRowData
	return []

'''
	Purpose: Deduplication on the provided csv file.
	Parameters: NONE.
	Return: NONE.
'''
def Deduplication():
	districtLowerRange = 101
	districtUpperRange = 206
	currentDistrictNumber = 0
	
	industrialCensus = GetCSVFile('industrial_header.csv', ',')
	
	# The single row (Containing all columns) of data in a CSV File.
	industrialCensusRowData = []
	
	# The matched data found in both files combined.
	duplicateRows = []
	
	foundMatch = False
	
	# Track the number of true matches found.
	industrialCensusRowCount = 0
	duplicateRowCount = 0
	percentageFound = 0.0
	
	# Start Timing.
	start_time = time.time()
	
	industrialCensusRowData = GetCSVFirstRow(industrialCensus)
	
	
	# Setup CSV Files.
	deduplicationWb = Workbook()
	deduplicationFilename = 'industrial_deduplication.xlsx'
	deduplicationWs = deduplicationWb.active
	deduplicationWs.title = "Industrial Deduplication"
	
	# Add data to the worksheet.
	tempList = []
	for row in industrialCensusRowData:
		tempList.append(row)
	deduplicationWs.append(tempList)	
	deduplicationWb.save(filename = deduplicationFilename)
	
	# Show results in new excel workbook.
	duplicatesWb = Workbook()
	duplicatesFilename = 'industrial_duplicates.xlsx'
	duplicatesWs = duplicatesWb.active
	duplicatesWs.title = "Industrial Duplicates"
	
	# Add data to the worksheet, some reason it doesn't work normally so had to do it this way.
	tempList = []
	for row in industrialCensusRowData:
		tempList.append(row)
	duplicatesWs.append(tempList)	
	duplicatesWb.save(filename = duplicatesFilename)
	
	for currentDistrictNumber in range(districtLowerRange, districtUpperRange+1):
		duplicateRows = []
		industrialCensusFilteredDistrict = DistrictFilter('industrial.csv', ',', 29, currentDistrictNumber)
		tempCSV = industrialCensusFilteredDistrict
		
		# Count Total rows.
		industrialCensusRowCount += GetCSVTotalRows(industrialCensusFilteredDistrict)
		
		# Begin the first pass.
		# Check if names have exact match.
		for rowIndex, row in enumerate(industrialCensusFilteredDistrict):
			#print ("Row Index: %s" % rowIndex)				
			foundDuplicate = False
			
			# Grab the name from the row.
			if (rowIndex != 0):
				proprior = row[8]
				proprior = proprior.split()
				duplicate = FindDuplicate(rowIndex, proprior, tempCSV)
				if duplicate:
					print ("Found Duplicate: %s" % duplicate)
					foundDuplicate = True
					
					#Collect the row data.
					duplicateRows.append(duplicate)
					duplicateRowCount = duplicateRowCount + 1
				else:
					foundDuplicate = False
			
				while (foundDuplicate):
					duplicate = FindDuplicate(rowIndex, proprior, tempCSV)
					if duplicate:
						print ("Found Duplicate: %s" % duplicate)
						foundDuplicate = True
						
						#Collect the row data.
						duplicateRows.append(duplicate)
						duplicateRowCount = duplicateRowCount + 1
					else:
						foundDuplicate = False
						
		# Remove all empty proprior
		for rowIndex, row in reversed(list(enumerate(tempCSV))):
			if (row[8] == []):
				#Collect the row data.
				duplicateRows.append(row)
				duplicateRowCount = duplicateRowCount + 1
				tempCSV.pop(rowIndex)
		
		for rowIndex, row in enumerate(duplicateRows):
				duplicatesWs.append(row)		
		duplicatesWb.save(filename = duplicatesFilename)
		
		for rowIndex, row in enumerate(tempCSV):
			if (rowIndex > 0):
				deduplicationWs.append(row)		
		deduplicationWb.save(filename = deduplicationFilename)	
		CompilationProgress(currentDistrictNumber, districtUpperRange)
		
	PercentageDuplicate(duplicateRowCount, industrialCensusRowCount)
	
	
	
	# Display time taken to user.			
	print("\nFunction took %s seconds to complete!" % (time.time() - start_time))
	
'''
	Purpose: Gets the record that matches fullname.
	Parameters: fullname (record to match to), csvFile (the file name for the csv file), industrialIDs (ids being used).
	Return: matched row.
'''
def FirstpassFindExactMatch(fullname, csvFile, industrialIDs):
	# dist : 8

	foundMatch = False
	matchedRowData = []
	
	tempFile = csvFile
		
	for rowIndex, row in enumerate(tempFile):
		#print ("\nRow: %s" % row)	
		if (rowIndex != 0 and row[41] not in industrialIDs):
			proprior = row[8]
			proprior = proprior.split()
			#print ("\nIndex: %s" % rowIndex)
			#print ("Big Census Fullname: %s" % fullname)
			
			if (proprior == fullname and (proprior != [] or fullname != [])):
				print ("\nExact Match")
				print ("Big Census Fullname: %s" % fullname)
				print ("Proprior: %s --- %s" % (proprior, row[41]))
				foundMatch = True
	
			if (foundMatch):
				matchedRowData = []
				for columnIndex, cell in enumerate(row):
					matchedRowData.append(cell)
					
				#print ("Found Match: %s" % matchedRowData)
				#print ("\nIndex: %s" % rowIndex)		
				#print ("FOUND Proprior: %s" % proprior)
				# Remove found record.
				#csvFile.pop(rowIndex)
				# Grab the ID field.
				industrialIDs.append(row[41])
				
				return matchedRowData
	return []
	
'''
	Purpose: Gets the record that matches fullname.
	Parameters: fullname (record to match to), csvFile (the file name for the csv file), industrialIDs (ids being used).
	Return: matched row.
'''
def SecondpassFindExactMatch(fullname, csvFile, industrialIDs):
	# dist : 8

	#print industrialIDs
	
	foundMatch = False
	matchedRowData = []
	
	tempFile = csvFile
		
	for rowIndex, row in enumerate(tempFile):
		#print ("\nRow: %s" % row)	
		if (rowIndex != 0 and row[41] not in industrialIDs):
			rowData = row[8]
			firstCheck = RemoveSpecificCharacters(rowData)
			firstCheck = ReplaceForwardSlashWithSpace(firstCheck)
			firstCheck = firstCheck.split()
			#print ("\nIndex: %s" % rowIndex)
			#print ("Base: %s" % rowData)
			#print ("Fullname: %s" % fullname)
			#print ("FirstCheck: %s" % firstCheck)
			
			fullnameTwoWords = []	
			# Ignore everything after first two words
			# Get only the lastname and firstname of fullname
			for i in range(len(fullname)):
				#print fullname[i]
				if  fullname[i]:
					if (i <= 1):
						fullnameTwoWords.append(fullname[i])
						#print secondCheck
			
			secondCheck = []
			# Ignore everything after first two words
			# Get only the lastname and firstname
			for i in range(len(firstCheck)):
				#print firstCheck[i]
				if  firstCheck[i]:
					if (i <= 1):
						secondCheck.append(firstCheck[i])
						#print secondCheck
			
			proprior = rowData
		
			if (firstCheck == fullname):
				print ("\nMatch - 1")
				print ("Big Census Fullname: %s" % fullname)
				print ("Proprior: %s" % firstCheck)
				foundMatch = True
			elif (secondCheck == fullname):
				print ("\nMatch - 2")
				print ("Big Census Fullname: %s" % fullname)
				print ("Proprior: %s" % firstCheck)
				foundMatch = True
			elif (firstCheck and fullnameTwoWords and firstCheck == fullnameTwoWords):
				print ("\nMatch - 3")
				print ("Big Census Fullname: %s" % fullname)
				print ("Proprior: %s" % firstCheck)
				foundMatch = True
			elif (secondCheck and fullnameTwoWords and secondCheck == fullnameTwoWords):
				print ("\nMatch - 4")
				print ("Big Census Fullname: %s" % fullname)
				print ("Proprior: %s" % secondCheck)
				foundMatch = True
	
			if (foundMatch):
				matchedRowData = []
				for columnIndex, cell in enumerate(row):
					matchedRowData.append(cell)
					
				# Check if there will be a comment added.
				comment = ""
				if "JR" in rowData:
					comment += " | JR"
				if "&" in rowData:
					comment += " | &"
				if "/" in rowData:
					comment += " | /"
				if "?" in rowData:
					comment += " | ?"
				if "," in rowData:
					comment += " | ,"
				# Remove first appearance of " / "
				comment = comment.replace(" | ", "", 1)
				
				matchedRowData.append(comment)
					
				#print ("Found Match: %s" % matchedRowData)
				#print ("\nIndex: %s" % rowIndex)	
				#print ("Big Census Fullname: %s" % fullname)
				#print ("FOUND Proprior: %s" % proprior)
				# Grab the ID field.
				industrialIDs.append(row[41])
				
				return matchedRowData
	return []
	
'''
	Purpose: Gets the record that matches fullname.
	Parameters: fullname (record to match to), sex (gender of fullname), csvFile (the file name for the csv file), industrialIDs (ids being used).
	Return: matched row.
'''
def FindPossibleMatch(fullname, sex, csvFile, industrialIDs):
	# dist : 8
	# SEX_CODE_ORIG : 44
	# SEX : 31

	#print industrialIDs
	
	
	foundMatch = False
	matchedRowData = []
	
	tempFile = csvFile
		
	for rowIndex, row in enumerate(tempFile):
		#print ("\nRow: %s" % row)	
		if (rowIndex != 0 and row[41] not in industrialIDs):
			rowData = row[8]
			firstCheck = RemoveSpecificCharacters(rowData)
			firstCheck = ReplaceForwardSlashWithSpace(firstCheck)
			firstCheck = firstCheck.split()
			#print ("\nIndex: %s" % rowIndex)
			#print ("Base: %s" % rowData)
			#print ("Fullname: %s" % fullname)
			#print ("FirstCheck: %s" % firstCheck)
					
			firstPossibleCheck = []
			secondPossibleCheck = []	
			thirdPossibleCheck = []
			
			performFirstPossibleCheck = False
			performSecondPossibleCheck = False
			performThirdPossibleCheck = False
			
			# Ignore everything after first two words
			# Get only the lastname and firstname with only first letter
			for i in range(len(firstCheck)):
				#print firstCheck[i]
				if  firstCheck[i]:
					if (i <= 1):
						if (i == 0):
							firstPossibleCheck.append(firstCheck[i])
							secondPossibleCheck.append(firstCheck[i])
							thirdPossibleCheck.append(firstCheck[i])
						elif (i == 1):
							tempWord = firstCheck[i]
							firstLetter = tempWord[:1]
							firstPossibleCheck.append(firstLetter)
							performFirstPossibleCheck = True
						elif (i == 2):
							tempWord = firstCheck[i]
							firstLetter = tempWord[:1]
							secondPossibleCheck.append(firstLetter)
							performSecondPossibleCheck = True
						elif (i == 3):
							tempWord = firstCheck[i]
							firstLetter = tempWord[:1]
							thirdPossibleCheck.append(firstLetter)
							performThirdPossibleCheck = True
			
			#print ("First Possible Check: %s" % firstPossibleCheck)
			#print ("Second Possible Check: %s" % secondPossibleCheck)
			#print ("Third Possible Check: %s" % thirdPossibleCheck)
					
			firstPossibleCheckFullname = []
			secondPossibleCheckFullname = []	
			thirdPossibleCheckFullname = []	
			
			performFirstPossibleCheckFullname = False
			performSecondPossibleCheckFullname = False
			performThirdPossibleCheckFullname = False
			
			# Ignore everything after first two words
			# Get only the lastname and firstname with only first letter
			for i in range(len(fullname)):
				#print fullname[i]
				if  fullname[i]:
					if (i <= 1):
						if (i == 0):
							firstPossibleCheckFullname.append(fullname[i])
							secondPossibleCheckFullname.append(fullname[i])
							thirdPossibleCheckFullname.append(fullname[i])
						elif (i == 1):
							tempWord = fullname[i]
							firstLetter = tempWord[:1]
							firstPossibleCheckFullname.append(firstLetter)
							performFirstPossibleCheckFullname = True
						elif (i == 2):
							tempWord = fullname[i]
							firstLetter = tempWord[:1]
							secondPossibleCheckFullname.append(firstLetter)
							performSecondPossibleCheckFullname = True
						elif (i == 3):
							tempWord = fullname[i]
							firstLetter = tempWord[:1]
							thirdPossibleCheckFullname.append(firstLetter)
							performThirdPossibleCheckFullname = True
			
			proprior = rowData
			
			#print ("Possible Check: %s" % firstPossibleCheck)
			#print ("Possible Check Fullname: %s" % firstPossibleCheckFullname)
		
			if (firstPossibleCheck == firstPossibleCheckFullname and (firstPossibleCheck != [] or firstPossibleCheckFullname != []) and (performFirstPossibleCheck and performFirstPossibleCheckFullname)):
				# Check sex.
				#print ("Big Census Sex: %s" % sex)
				#print ("Proprior Sex: %s" % row[31])
				if (row[31] == "0" and sex == "M"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 1 - Male")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
				elif (row[31] == "1" and sex == "F"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 1 - Female")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
				elif (row[31] == "2"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 1 - Male+Female")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
			elif (secondPossibleCheck == secondPossibleCheckFullname and (secondPossibleCheck != [] or secondPossibleCheckFullname != []) and (performSecondPossibleCheck and performSecondPossibleCheckFullname)):
				# Check sex.
				#print ("Big Census Sex: %s" % sex)
				#print ("Proprior Sex: %s" % row[31])
				if (row[31] == "0" and sex == "M"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 2 - Male")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
				elif (row[31] == "1" and sex == "F"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 2 - Female")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
				elif (row[31] == "2"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 2 - Male+Female")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
			elif (thirdPossibleCheck == thirdPossibleCheckFullname and (thirdPossibleCheck != [] or thirdPossibleCheckFullname != []) and (performThirdPossibleCheck and performThirdPossibleCheckFullname)):
				# Check sex.
				#print ("Big Census Sex: %s" % sex)
				#print ("Proprior Sex: %s" % row[31])
				if (row[31] == "0" and sex == "M"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 3 - Male")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
				elif (row[31] == "1" and sex == "F"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 3 - Female")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
				elif (row[31] == "2"):
					# Makes sure not to grab any empty names that may still be in the census data.
					print ("\nPossible Match 3 - Male+Female")
					print ("Big Census Fullname: %s" % fullname)
					print ("Proprior: %s" % firstCheck)
					foundMatch = True
	
			if (foundMatch):
				matchedRowData = []
				for columnIndex, cell in enumerate(row):
					matchedRowData.append(cell)
					
				# Check if there will be a comment added.
				comment = ""
				if "JR" in rowData:
					comment += " | JR"
				if "&" in rowData:
					comment += " | &"
				if "/" in rowData:
					comment += " | /"
				if "?" in rowData:
					comment += " | ?"
				if "," in rowData:
					comment += " | ,"
				# Remove first appearance of " / "
				comment = comment.replace(" | ", "", 1)
				
				matchedRowData.append(comment)
					
				#print ("Found Match: %s" % matchedRowData)
				#print ("\nIndex: %s" % rowIndex)	
				#print ("Big Census Fullname: %s" % fullname)
				#print ("FOUND Proprior: %s" % proprior)
				# Grab the ID field.
				industrialIDs.append(row[41])
				
				return matchedRowData
	return []
	
'''
	Purpose: Remove a few special characters from provided string.
	Parameters: string (string to remove special characters from).
	Return: string results.
'''
def RemoveSpecificCharacters(string):
	string = re.sub('[,?]', '', string)
	return string
	
'''
	Purpose: Change a few special characters from provided string.
	Parameters: string (string to change special characters from).
	Return: string results.
'''
def ReplaceForwardSlashWithSpace(string):
	string = re.sub('[/]', ' ', string)
	return string
	
'''
	Purpose: Display the amount matched based on district.
	Parameters: matchedRowCount (Matched row count total), industrialCensusRowCount ( the total rows in industrial census for district).
	Return: NONE.
'''
def PercentageMatched(matchedRowCount, industrialCensusRowCount):
	percentageMatched = 0.0
	if (matchedRowCount > 0):
		percentageMatched = float(matchedRowCount) / float(industrialCensusRowCount)
	recordsNotMatched = industrialCensusRowCount - matchedRowCount
	print("\n***********************************")
	print("Industrial Census Record Amount: %d" % industrialCensusRowCount)
	print("True Match Record Amount: %d" % matchedRowCount)
	print("Remaining Records Not Matched: %s" % recordsNotMatched)
	print("Percentage of Records Matched: %.2f%%" % percentageMatched)
	print("***********************************")
	
'''
	Purpose: Display the duplicate rows.
	Parameters: duplicateRowCount (Matched duplicate row count total), industrialCensusRowCount ( the total rows in industrial census for district).
	Return: NONE.
'''
def PercentageDuplicate(duplicateRowCount, industrialCensusRowCount):
	percentageDuplicate = 0.0
	if (duplicateRowCount > 0):
		percentageDuplicate = float(duplicateRowCount) / float(industrialCensusRowCount)
	recordsNotDuplicate = industrialCensusRowCount - duplicateRowCount
	print("\n***********************************")
	print("Industrial Census Record Amount: %d" % industrialCensusRowCount)
	print("Duplicate Record Amount: %d" % duplicateRowCount)
	print("Unique Record Amount: %s" % recordsNotDuplicate)
	print("Percentage of Records that are Duplicates: %.2f%%" % percentageDuplicate)
	print("***********************************")

'''
	Purpose: Display the current progress of the record linkage state.
	Parameters: currentDistrictNumber (Current district that was just finished), districtUpperRange (Last district to perform record linkage on).
	Return: NONE.
'''
def CompilationProgress(currentDistrictNumber, districtUpperRange):
	compilationProgress = float(currentDistrictNumber) / float(districtUpperRange)
	print("\n***********************************")
	print("Current District Number: %s" % currentDistrictNumber)
	print("Last District Number: %s" % districtUpperRange)
	print("Compilation Progress: %.3f%%" % compilationProgress)
	print("***********************************")	
	
def main():
	userInput = '0'
	checker = 1

	menu()

	while (checker == 1):
		userInput = raw_input("\nPlease enter a menu option: ")
		if (userInput == '1'):
			print ("\n\n First Pass.\n\n")
			
			districtLowerRange = 1
			districtUpperRange = 50
			currentDistrictNumber = 0

			# Get the header data to be used at the first row of the new csv file.
			#industrialCensus = GetCSVFile('industrial_header.csv', ',')
			industrialCensus = [ 	"ced", "empboy", "empgirl", "estcode", "procode",
									"subref", "month", "csd", "proprior", "typeest",
									"sic", "typepow", "force", "fixcap", "flocap",
									"empmen", "empwom", "wages", "refnum", "comments",
									"PROD1", "PROD2", "PROD3", "RAWMAT1", "RAWMAT2",
									"RAWMAT3", "sumproc", "sumrawc", "ind", "dist",
									"district", "sex", "power", "cedref", "comcode",
									"jntcode", "IND2", "IND3", "ref", "pow",
									"prov", "id", "record comment"
								]
			#fullCensus = GetCSVFile('1871_header.csv', ',')
			fellCensus		 = [ 	"FOLDER_IMAGE_SEQ", "IMAGE_ID", "FOLDER", "AGE_IN_YEARS", "BATCH_ID",
									"BATCH_LOCALITY", "BIRTH_DATE_STD", "BIRTH_PLACE", "CENSUS_PLACE", "DIGITAL_GS_NUMBER",
									"DISTRICT_NAME", "DISTRICT_NUMBER", "DIVISION_NUMBER", "DWELLING", "EASY_IMAGE_ID",
									"ESTIMATED_BIRTH_YEAR", "EVENT_DATE_STD", "FILM_NUMBER", "HOUSEHOLD_ID", "IMAGE_NBR",
									"IMAGE_TYPE", "LAC_FILM_NUMBER", "LINE_NBR", "MARITAL_STATUS", "MARITAL_STATUS_ORIG",
									"NUMERIC_IMAGE_NBR", "ORIGIN", "PACKET_LTR", "PADDED_DGS_NUMBER", "PAGE",
									"PR_AGE_ORIG", "PR_NAME", "PR_NAME_GN", "PR_NAME_GN_ORIG", "PR_NAME_ORIG",
									"PR_NAME_SURN", "PR_NAME_SURN_ORIG", "PROVINCE", "REC_NBR", "RECORD_ID",
									"RECORD_GROUP", "RECORD_PLACE", "RELIGION", "RESIDENCE", "SEX_CODE_ORIG",
									"SEX_CODE_STD", "SUB_DISTRICT_NAME", "SUB_DISTRICT_NUMBER", "UDE_BATCH_NUMBER", "UNIQUE_IDENTIFIER"
								]
			
			# The single row (Containing all columns) of data in a CSV File.
			industrialCensusRowData = []
			fullCensusRowData = []
			
			# The matched data found in both files combined.
			matchedRows = []
			
			foundMatch = False
			
			# Track the number of true matches found.
			industrialCensusRowCount = 0
			matchedRowCount = 0
			percentageFound = 0.0
			
			# Start Timing.
			start_time = time.time()
			
			# Get First Row data of both CSV files.
			fullCensusRowData = fellCensus
			industrialCensusRowData = industrialCensus
			# Combine both the header data.
			combinedList = fullCensusRowData + ["match"] + industrialCensusRowData
			matchedRows.append(combinedList)
			
			# PR_NAME_GN : 32
			# PR_NAME_GN_ORIG : 33
			# PR_NAME_ORIG : 34
			# PR_NAME_SURN : 35
			# PR_NAME_SURN_ORIG : 36
			
			# Setup CSV Files.
			firstpassMatchesWb = Workbook()
			firstpassMatchesfilename = 'firstpass_matches.xlsx'
			firstpassMatchesWs = firstpassMatchesWb.active
			firstpassMatchesWs.title = "One-to-one"
			firstpassMatchesWs2 = firstpassMatchesWb.create_sheet(title="Many")
			
			# Add data to the worksheet.
			for row in matchedRows:
				firstpassMatchesWs.append(row)
				firstpassMatchesWs2.append(row)
			firstpassMatchesWb.save(filename = firstpassMatchesfilename)
			
			# Show results in new excel workbook.
			firstpassNoMatchesWb = Workbook()
			firstpassNoMatchesfilename = 'firstpass_nomatches.xlsx'
			firstpassNoMatchesWs = firstpassNoMatchesWb.active
			firstpassNoMatchesWs.title = "No Matches"
			
			# Add data to the worksheet, some reason it doesn't work normally so had to do it this way.
			tempList = []
			for row in industrialCensusRowData:
				tempList.append(row)
			firstpassNoMatchesWs.append(tempList)	
			firstpassNoMatchesWb.save(filename = firstpassNoMatchesfilename)
			
			for currentDistrictNumber in range(districtLowerRange, districtUpperRange+1):
				matchedRows = []
				fullCensusFilteredDistrict = DistrictFilter('1871full.txt', '|', 11, currentDistrictNumber)
				industrialCensusFilteredDistrict = DistrictFilter('industrial.csv', ',', 29, currentDistrictNumber)
				
				# Temp census for duplicate record checking.
				tempIndustrialCensusFilteredDistrict = industrialCensusFilteredDistrict
				tempCensus = fullCensusFilteredDistrict
				#print tempCensus
				
				# Count Total rows.
				industrialCensusRowCount += GetCSVTotalRows(industrialCensusFilteredDistrict)
				
				# Begin the first pass.
				# Check if names have exact match.
				for rowIndex, row in enumerate(fullCensusFilteredDistrict):
					#print ("Row Index: %s" % rowIndex)				
					foundMatch = False
					count = 0
					# Grab the name from the row.
					if (rowIndex != 0):
						industrialCensusRowData = []
						industrialIDs = []
						fullname = []
						firstname = row[32]
						lastname = row[35]
						# Split the last name, special to second pass.
						firstname = firstname.upper()
						firstnameList = []
						firstnameList = firstname.split()
						#print lastname
						#print lastnameList
								
						fullname.append(lastname.upper())
						fullname = fullname + firstnameList
						firstMatch = True
						foundDuplicate = False
						match = FirstpassFindExactMatch(fullname, industrialCensusFilteredDistrict, industrialIDs)
						while match:
							#print match
							industrialCensusRowData.append(match)
							match = FirstpassFindExactMatch(fullname, industrialCensusFilteredDistrict, industrialIDs)
							foundMatch = True
							
							count = count + 1
							
							if not match:
								if (count >= 2):
									print ("---------------------------------------------> MULTIPLE")
								print ("\n*****************************")
					
					if (foundMatch):
						fullCensusRowData = []
						for columnIndex, cell in enumerate(row):
							fullCensusRowData.append(cell)
						
						# Collect the row data.
						for multipleRowIndex, multipleRow in enumerate(industrialCensusRowData):
							combinedList = fullCensusRowData + ["< ===== >"] + multipleRow
							matchedRows.append(combinedList)

				# Add the Matched Rows data to the worksheet.
				# Sort based on Census field name "PR_NAME" so same names are grouped together.
				matchedRows.sort(key=lambda x:x[31])
				
				tempCSV = industrialCensusFilteredDistrict
				# Check in the list which records were not matched.
				for rowIndex, row in enumerate(matchedRows):
					# INDUSTRIAL ID COMBINED = 92
					# INDUSTRIAL ID = 41
					# Find if it matched to anything in the industrial census and remove it.
					for industRowIndex, industRow in reversed(list(enumerate(tempCSV))):
						if (row[92] == industRow[41]):
							#print "FOUND MATCH"
							industrialCensusFilteredDistrict.pop(industRowIndex)
							matchedRowCount = matchedRowCount + 1
					
				
				# Check in the list which is one-to-one or a many relationship and split them up.
				oneToOneList = []
				manyList = []
				tempList = []
				counter = 0
				
				# COMBINED IDs
				# CENSUS NAME = 31
				# CENSUS ID = 49
				# INDUSTRIAL NAME = 59
				# INDUSTRIAL ID = 92
				
				tempCensusId = -1
				tempIndustrialId = -1
				for rowIndex, row in enumerate(matchedRows):	
					#print("%s, %s -- %s, %s" % (row[31], row[49], row[59], row[92]))
					#print("\nCENSUS ID: %s " % row[49])
					#print("\nINDUSTRIAL NAME: %s " % row[59])
					#print("\nINDUSTRIAL ID: %s " % row[92])
					if (rowIndex == 0):
						tempCensusId = row[49]
						tempIndustrialId = row[92]
					
					if (tempCensusId != row[49] and tempIndustrialId != row[92]):
						# Found a new type of row.
						if (counter <= 1):
							oneToOneList = oneToOneList + tempList
						else:
							manyList = manyList + tempList
							
						counter = 0
						tempList = []
						#oneToOneList.append(row)
					counter = counter + 1
					tempList.append(row)
						
					tempCensusId = row[49]
					tempIndustrialId = row[92]
				# Get the last one also.
				if (counter <= 1):
					oneToOneList = oneToOneList + tempList
				else:
					manyList = manyList + tempList
					
					
				# Write to excel.
				for rowIndex, row in enumerate(oneToOneList):
						# One-to-one sheet.
						firstpassMatchesWs.append(row)
				for rowIndex, row in enumerate(manyList):
						#firstpassMatchesWs.append(row)
						# Many sheet.
						firstpassMatchesWs2.append(row)
				firstpassMatchesWb.save(filename = firstpassMatchesfilename)
				
				# Add the Not Matched Rows data to the worksheet.
				for rowIndex, row in enumerate(industrialCensusFilteredDistrict):
					if (rowIndex > 0):
						firstpassNoMatchesWs.append(row)		
				firstpassNoMatchesWb.save(filename = firstpassNoMatchesfilename)
				
				CompilationProgress(currentDistrictNumber, districtUpperRange)
				
			PercentageMatched(matchedRowCount, industrialCensusRowCount)
			
			# Display time taken to user.			
			print("\nFunction took %s seconds to complete!" % (time.time() - start_time))
			
			# Display menu to user again.
			menu()
		elif (userInput == '2'):
			print ("\n\n Second Pass.\n\n")

			districtLowerRange = 1
			districtUpperRange = 3
			currentDistrictNumber = 0

			# Get the header data to be used at the first row of the new csv file.
			#industrialCensus = GetCSVFile('industrial_header.csv', ',')
			industrialCensus = [ 	"ced", "empboy", "empgirl", "estcode", "procode",
									"subref", "month", "csd", "proprior", "typeest",
									"sic", "typepow", "force", "fixcap", "flocap",
									"empmen", "empwom", "wages", "refnum", "comments",
									"PROD1", "PROD2", "PROD3", "RAWMAT1", "RAWMAT2",
									"RAWMAT3", "sumproc", "sumrawc", "ind", "dist",
									"district", "sex", "power", "cedref", "comcode",
									"jntcode", "IND2", "IND3", "ref", "pow",
									"prov", "id", "record comment"
								]
			#fullCensus = GetCSVFile('1871_header.csv', ',')
			fellCensus		 = [ 	"FOLDER_IMAGE_SEQ", "IMAGE_ID", "FOLDER", "AGE_IN_YEARS", "BATCH_ID",
									"BATCH_LOCALITY", "BIRTH_DATE_STD", "BIRTH_PLACE", "CENSUS_PLACE", "DIGITAL_GS_NUMBER",
									"DISTRICT_NAME", "DISTRICT_NUMBER", "DIVISION_NUMBER", "DWELLING", "EASY_IMAGE_ID",
									"ESTIMATED_BIRTH_YEAR", "EVENT_DATE_STD", "FILM_NUMBER", "HOUSEHOLD_ID", "IMAGE_NBR",
									"IMAGE_TYPE", "LAC_FILM_NUMBER", "LINE_NBR", "MARITAL_STATUS", "MARITAL_STATUS_ORIG",
									"NUMERIC_IMAGE_NBR", "ORIGIN", "PACKET_LTR", "PADDED_DGS_NUMBER", "PAGE",
									"PR_AGE_ORIG", "PR_NAME", "PR_NAME_GN", "PR_NAME_GN_ORIG", "PR_NAME_ORIG",
									"PR_NAME_SURN", "PR_NAME_SURN_ORIG", "PROVINCE", "REC_NBR", "RECORD_ID",
									"RECORD_GROUP", "RECORD_PLACE", "RELIGION", "RESIDENCE", "SEX_CODE_ORIG",
									"SEX_CODE_STD", "SUB_DISTRICT_NAME", "SUB_DISTRICT_NUMBER", "UDE_BATCH_NUMBER", "UNIQUE_IDENTIFIER"
								]
			
			# The single row (Containing all columns) of data in a CSV File.
			industrialCensusRowData = []
			fullCensusRowData = []
			
			# The matched data found in both files combined.
			matchedRows = []
			
			foundMatch = False
			
			# Track the number of true matches found.
			industrialCensusRowCount = 0
			matchedRowCount = 0
			percentageFound = 0.0
			
			# Start Timing.
			start_time = time.time()
			
			# Get First Row data of both CSV files.
			fullCensusRowData = fellCensus
			industrialCensusRowData = industrialCensus
			# Combine both the header data.
			combinedList = fullCensusRowData + ["match"] + industrialCensusRowData
			matchedRows.append(combinedList)
			
			# PR_NAME_GN : 32
			# PR_NAME_GN_ORIG : 33
			# PR_NAME_ORIG : 34
			# PR_NAME_SURN : 35
			# PR_NAME_SURN_ORIG : 36
			
			# Setup CSV Files.
			firstpassMatchesWb = Workbook()
			firstpassMatchesfilename = 'secondpass_matches.xlsx'
			firstpassMatchesWs = firstpassMatchesWb.active
			firstpassMatchesWs.title = "One-to-one"
			firstpassMatchesWs2 = firstpassMatchesWb.create_sheet(title="Many")
			
			# Add data to the worksheet.
			for row in matchedRows:
				firstpassMatchesWs.append(row)
				firstpassMatchesWs2.append(row)
			firstpassMatchesWb.save(filename = firstpassMatchesfilename)
			
			# Show results in new excel workbook.
			firstpassNoMatchesWb = Workbook()
			firstpassNoMatchesfilename = 'secondpass_nomatches.xlsx'
			firstpassNoMatchesWs = firstpassNoMatchesWb.active
			firstpassNoMatchesWs.title = "No Matches"
			
			# Add data to the worksheet, some reason it doesn't work normally so had to do it this way.
			tempList = []
			for row in industrialCensusRowData:
				tempList.append(row)
			firstpassNoMatchesWs.append(tempList)	
			firstpassNoMatchesWb.save(filename = firstpassNoMatchesfilename)
			
			for currentDistrictNumber in range(districtLowerRange, districtUpperRange+1):
				matchedRows = []
				fullCensusFilteredDistrict = DistrictFilter('1871full.txt', '|', 11, currentDistrictNumber)
				industrialCensusFilteredDistrict = DistrictFilter('firstpass_nomatches.csv', ',', 29, currentDistrictNumber)
				
				# Temp census for duplicate record checking.
				tempIndustrialCensusFilteredDistrict = industrialCensusFilteredDistrict
				tempCensus = fullCensusFilteredDistrict
				#print tempCensus
				
				# Count Total rows.
				industrialCensusRowCount += GetCSVTotalRows(industrialCensusFilteredDistrict)
				
				# Begin the first pass.
				# Check if names have exact match.
				for rowIndex, row in enumerate(fullCensusFilteredDistrict):
					#print ("Row Index: %s" % rowIndex)				
					foundMatch = False
					
					# Grab the name from the row.
					if (rowIndex != 0):
						industrialIDs = []
						fullname = []
						firstname = row[32]
						lastname = row[35]
						# Split the last name, special to second pass.
						firstname = firstname.upper()
						firstnameList = []
						firstnameList = firstname.split()
						#print lastname
						#print lastnameList
								
						fullname.append(lastname.upper())
						fullname = fullname + firstnameList
						firstMatch = True
						foundDuplicate = False
						count = 0
						match = SecondpassFindExactMatch(fullname, industrialCensusFilteredDistrict, industrialIDs)
						while match:
							industrialCensusRowData = match
							match = SecondpassFindExactMatch(fullname, industrialCensusFilteredDistrict, industrialIDs)
							foundMatch = True
							
							count = count + 1
							
							if not match:
								if (count >= 2):
									print ("######################### Big Census Record many to one Industrial Census")
								print ("\n*****************************")
					
					if (foundMatch):
						fullCensusRowData = []
						for columnIndex, cell in enumerate(row):
							fullCensusRowData.append(cell)
						
						# Collect the row data.
						combinedList = fullCensusRowData + ["< ===== >"] + industrialCensusRowData
						matchedRows.append(combinedList)

				# Add the Matched Rows data to the worksheet.
				# Sort based on Census field name "PR_NAME" so same names are grouped together.
				matchedRows.sort(key=lambda x:x[31])
				
				tempCSV = industrialCensusFilteredDistrict
				# Check in the list which records were not matched.
				for rowIndex, row in enumerate(matchedRows):
					# INDUSTRIAL ID COMBINED = 92
					# INDUSTRIAL ID = 41
					# Find if it matched to anything in the industrial census and remove it.
					for industRowIndex, industRow in reversed(list(enumerate(tempCSV))):
						if (row[92] == industRow[41]):
							#print "FOUND MATCH"
							industrialCensusFilteredDistrict.pop(industRowIndex)
							matchedRowCount = matchedRowCount + 1
					
				
				# Check in the list which is one-to-one or a many relationship and split them up.
				oneToOneList = []
				manyList = []
				tempList = []
				counter = 0
				
				# COMBINED IDs
				# CENSUS NAME = 31
				# CENSUS ID = 49
				# INDUSTRIAL NAME = 59
				# INDUSTRIAL ID = 92
				
				tempCensusId = -1
				tempIndustrialId = -1
				for rowIndex, row in enumerate(matchedRows):	
					#print("%s, %s -- %s, %s" % (row[31], row[49], row[59], row[92]))
					#print("\nCENSUS ID: %s " % row[49])
					#print("\nINDUSTRIAL NAME: %s " % row[59])
					#print("\nINDUSTRIAL ID: %s " % row[92])
					if (rowIndex == 0):
						tempCensusId = row[49]
						tempIndustrialId = row[92]
					
					if (tempCensusId != row[49] and tempIndustrialId != row[92]):
						# Found a new type of row.
						if (counter <= 1):
							oneToOneList = oneToOneList + tempList
						else:
							manyList = manyList + tempList
							
						counter = 0
						tempList = []
						#oneToOneList.append(row)
					counter = counter + 1
					tempList.append(row)
						
					tempCensusId = row[49]
					tempIndustrialId = row[92]
				# Get the last one also.
				if (counter <= 1):
					oneToOneList = oneToOneList + tempList
				else:
					manyList = manyList + tempList
					
					
				# Write to excel.
				for rowIndex, row in enumerate(oneToOneList):
						# One-to-one sheet.
						firstpassMatchesWs.append(row)
				for rowIndex, row in enumerate(manyList):
						#firstpassMatchesWs.append(row)
						# Many sheet.
						firstpassMatchesWs2.append(row)
				firstpassMatchesWb.save(filename = firstpassMatchesfilename)
				
				# Add the Not Matched Rows data to the worksheet.
				for rowIndex, row in enumerate(industrialCensusFilteredDistrict):
					if (rowIndex > 0):
						firstpassNoMatchesWs.append(row)		
				firstpassNoMatchesWb.save(filename = firstpassNoMatchesfilename)
				
				CompilationProgress(currentDistrictNumber, districtUpperRange)
				
			PercentageMatched(matchedRowCount, industrialCensusRowCount)
			
			# Display time taken to user.			
			print("\nFunction took %s seconds to complete!" % (time.time() - start_time))
			
			# Display menu to user again.
			menu()
		elif (userInput == '3'):
			print ("\n\n Find Possible Matches.\n\n")

			districtLowerRange = 151
			districtUpperRange = 206
			currentDistrictNumber = 0

			# Get the header data to be used at the first row of the new csv file.
			#industrialCensus = GetCSVFile('industrial_header.csv', ',')
			industrialCensus = [ 	"ced", "empboy", "empgirl", "estcode", "procode",
									"subref", "month", "csd", "proprior", "typeest",
									"sic", "typepow", "force", "fixcap", "flocap",
									"empmen", "empwom", "wages", "refnum", "comments",
									"PROD1", "PROD2", "PROD3", "RAWMAT1", "RAWMAT2",
									"RAWMAT3", "sumproc", "sumrawc", "ind", "dist",
									"district", "sex", "power", "cedref", "comcode",
									"jntcode", "IND2", "IND3", "ref", "pow",
									"prov", "id", "record comment"
								]
			#fullCensus = GetCSVFile('1871_header.csv', ',')
			fellCensus		 = [ 	"FOLDER_IMAGE_SEQ", "IMAGE_ID", "FOLDER", "AGE_IN_YEARS", "BATCH_ID",
									"BATCH_LOCALITY", "BIRTH_DATE_STD", "BIRTH_PLACE", "CENSUS_PLACE", "DIGITAL_GS_NUMBER",
									"DISTRICT_NAME", "DISTRICT_NUMBER", "DIVISION_NUMBER", "DWELLING", "EASY_IMAGE_ID",
									"ESTIMATED_BIRTH_YEAR", "EVENT_DATE_STD", "FILM_NUMBER", "HOUSEHOLD_ID", "IMAGE_NBR",
									"IMAGE_TYPE", "LAC_FILM_NUMBER", "LINE_NBR", "MARITAL_STATUS", "MARITAL_STATUS_ORIG",
									"NUMERIC_IMAGE_NBR", "ORIGIN", "PACKET_LTR", "PADDED_DGS_NUMBER", "PAGE",
									"PR_AGE_ORIG", "PR_NAME", "PR_NAME_GN", "PR_NAME_GN_ORIG", "PR_NAME_ORIG",
									"PR_NAME_SURN", "PR_NAME_SURN_ORIG", "PROVINCE", "REC_NBR", "RECORD_ID",
									"RECORD_GROUP", "RECORD_PLACE", "RELIGION", "RESIDENCE", "SEX_CODE_ORIG",
									"SEX_CODE_STD", "SUB_DISTRICT_NAME", "SUB_DISTRICT_NUMBER", "UDE_BATCH_NUMBER", "UNIQUE_IDENTIFIER"
								]
			
			# The single row (Containing all columns) of data in a CSV File.
			industrialCensusRowData = []
			fullCensusRowData = []
			
			# The matched data found in both files combined.
			matchedRows = []
			
			foundMatch = False
			
			# Track the number of true matches found.
			industrialCensusRowCount = 0
			matchedRowCount = 0
			percentageFound = 0.0
			
			# Start Timing.
			start_time = time.time()
			
			# Get First Row data of both CSV files.
			fullCensusRowData = fellCensus
			industrialCensusRowData = industrialCensus
			# Combine both the header data.
			combinedList = fullCensusRowData + ["match"] + industrialCensusRowData
			matchedRows.append(combinedList)
			
			# PR_NAME_GN : 32
			# PR_NAME_GN_ORIG : 33
			# PR_NAME_ORIG : 34
			# PR_NAME_SURN : 35
			# PR_NAME_SURN_ORIG : 36
			
			# Setup CSV Files.
			firstpassMatchesWb = Workbook()
			firstpassMatchesfilename = 'possible_matches.xlsx'
			firstpassMatchesWs = firstpassMatchesWb.active
			firstpassMatchesWs.title = "One-to-one"
			firstpassMatchesWs2 = firstpassMatchesWb.create_sheet(title="Many")
			
			# Add data to the worksheet.
			for row in matchedRows:
				firstpassMatchesWs.append(row)
				firstpassMatchesWs2.append(row)
			firstpassMatchesWb.save(filename = firstpassMatchesfilename)
			
			# Show results in new excel workbook.
			firstpassNoMatchesWb = Workbook()
			firstpassNoMatchesfilename = 'possible_nomatches.xlsx'
			firstpassNoMatchesWs = firstpassNoMatchesWb.active
			firstpassNoMatchesWs.title = "No Matches"
			
			# Add data to the worksheet, some reason it doesn't work normally so had to do it this way.
			tempList = []
			for row in industrialCensusRowData:
				tempList.append(row)
			firstpassNoMatchesWs.append(tempList)	
			firstpassNoMatchesWb.save(filename = firstpassNoMatchesfilename)
			
			for currentDistrictNumber in range(districtLowerRange, districtUpperRange+1):
				matchedRows = []
				fullCensusFilteredDistrict = DistrictFilter('1871full.txt', '|', 11, currentDistrictNumber)
				industrialCensusFilteredDistrict = DistrictFilter('secondpass_nomatches.csv', ',', 29, currentDistrictNumber)
				
				# Temp census for duplicate record checking.
				tempIndustrialCensusFilteredDistrict = industrialCensusFilteredDistrict
				tempCensus = fullCensusFilteredDistrict
				#print tempCensus
				
				# Count Total rows.
				industrialCensusRowCount += GetCSVTotalRows(industrialCensusFilteredDistrict)
				
				# Begin the first pass.
				# Check if names have exact match.
				for rowIndex, row in enumerate(fullCensusFilteredDistrict):
					#print ("Row Index: %s" % rowIndex)				
					foundMatch = False
					count = 0
					# Grab the name from the row.
					if (rowIndex != 0):
						industrialCensusRowData = []
						industrialIDs = []
						fullname = []
						firstname = row[32]
						lastname = row[35]
						# Split the last name, special to second pass.
						firstname = firstname.upper()
						firstnameList = []
						firstnameList = firstname.split()
						#print lastname
						#print lastnameList
								
						fullname.append(lastname.upper())
						fullname = fullname + firstnameList
						firstMatch = True
						foundDuplicate = False
						match = FindPossibleMatch(fullname, row[44], industrialCensusFilteredDistrict, industrialIDs)
						while match:
							industrialCensusRowData.append(match)
							match = FindPossibleMatch(fullname, row[44], industrialCensusFilteredDistrict, industrialIDs)
							foundMatch = True
							
							count = count + 1
							
							if not match:
								if (count >= 2):
									print ("---------------------------------------------> MULTIPLE")
								print ("\n*****************************")
					
					if (foundMatch):
						fullCensusRowData = []
						for columnIndex, cell in enumerate(row):
							fullCensusRowData.append(cell)
						
						# Collect the row data.
						for multipleRowIndex, multipleRow in enumerate(industrialCensusRowData):
							combinedList = fullCensusRowData + ["< ===== >"] + multipleRow
							matchedRows.append(combinedList)

				# Add the Matched Rows data to the worksheet.
				# Sort based on Census field name "PR_NAME" so same names are grouped together.
				matchedRows.sort(key=lambda x:x[31])
				
				tempCSV = industrialCensusFilteredDistrict
				# Check in the list which records were not matched.
				for rowIndex, row in enumerate(matchedRows):
					# INDUSTRIAL ID COMBINED = 92
					# INDUSTRIAL ID = 41
					# Find if it matched to anything in the industrial census and remove it.
					for industRowIndex, industRow in reversed(list(enumerate(tempCSV))):
						if (row[92] == industRow[41]):
							#print "FOUND MATCH"
							industrialCensusFilteredDistrict.pop(industRowIndex)
							matchedRowCount = matchedRowCount + 1
					
				
				# Check in the list which is one-to-one or a many relationship and split them up.
				bigCensusIdList = []
				industrialCensusIdList = []
				oneToOneList = []
				manyList = []
				tempList = []
				bigCounter = 0
				industrialCounter = 0
				
				# Collect all the IDs in the matched rows.
				for rowIndex, row in enumerate(matchedRows):
					bigCensusIdList.append(row[49])
					industrialCensusIdList.append(row[92])
					
				
				
				# COMBINED IDs
				# CENSUS NAME = 31
				# CENSUS ID = 49
				# INDUSTRIAL NAME = 59
				# INDUSTRIAL ID = 92
				
				tempBigCensusId = -1
				tempIndustrialId = -1
				for rowIndex, row in enumerate(matchedRows):	
					#print("%s, %s -- %s, %s" % (row[31], row[49], row[59], row[92]))
					#print("\nCENSUS ID: %s " % row[49])
					#print("\nINDUSTRIAL NAME: %s " % row[59])
					#print("\nINDUSTRIAL ID: %s " % row[92])
					tempBigCensusId = row[49]
					tempIndustrialId = row[92]
					
					bigCounter = bigCensusIdList.count(tempBigCensusId)
					industrialCounter = industrialCensusIdList.count(tempIndustrialId)
						
					if (bigCounter == 1 and industrialCounter == 1):
						#oneToOneList = oneToOneList + row
						oneToOneList.append(row)
					else:
						#manyList = manyList + row
						manyList.append(row)
					
				# Write to excel.
				for rowIndex, row in enumerate(oneToOneList):
						# One-to-one sheet.
						firstpassMatchesWs.append(row)
				for rowIndex, row in enumerate(manyList):
						#firstpassMatchesWs.append(row)
						# Many sheet.
						firstpassMatchesWs2.append(row)
				firstpassMatchesWb.save(filename = firstpassMatchesfilename)
				
				# Add the Not Matched Rows data to the worksheet.
				for rowIndex, row in enumerate(industrialCensusFilteredDistrict):
					if (rowIndex > 0):
						firstpassNoMatchesWs.append(row)		
				firstpassNoMatchesWb.save(filename = firstpassNoMatchesfilename)
				
				CompilationProgress(currentDistrictNumber, districtUpperRange)
				
			PercentageMatched(matchedRowCount, industrialCensusRowCount)
			
			# Display time taken to user.			
			print("\nFunction took %s seconds to complete!" % (time.time() - start_time))
			
			# Display menu to user again.
			menu()
		elif (userInput == '4' or userInput == 'q'):
			print("\nNow quitting the program!\n")
			checker = 0
		else:
			print("Incorrect input, try again!\n")
			menu()

if __name__ == "__main__":
	main()